from openpyxl import load_workbook
import os

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

resources_path = os.path.join(os.path.dirname(__file__), '..', 'resources')
excel_file_path = os.path.join(resources_path, 'file_example_XLSX_50.xlsx')


def test_excel_file():
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    cell_value = sheet.cell(row=3, column=2).value
    print(cell_value)

    assert cell_value == 'Mara'
    assert sheet.max_row == 51
